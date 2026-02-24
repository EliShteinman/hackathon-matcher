import io
import logging

from openpyxl import load_workbook
from pydantic import ValidationError

from config.excel_settings import ExcelSettings
from src.exceptions.excel_exceptions import InvalidExcelFormatError, MissingRequiredColumnError
from src.models.user import User
from src.repositories.interfaces.user_repository import UserRepositoryInterface
from src.schemas.excel_schemas import ExcelRowSchema

logger = logging.getLogger(__name__)


class ExcelService:
    def __init__(self, user_repository: UserRepositoryInterface, excel_settings: ExcelSettings) -> None:
        self._user_repo = user_repository
        self._settings = excel_settings

    async def import_from_bytes(self, file_content: bytes) -> int:
        try:
            wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise InvalidExcelFormatError(str(exc)) from exc

        if self._settings.sheet_name:
            if self._settings.sheet_name not in wb.sheetnames:
                raise InvalidExcelFormatError(f"Sheet '{self._settings.sheet_name}' not found")
            ws = wb[self._settings.sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(min_row=self._settings.header_row, values_only=False))
        if not rows:
            raise InvalidExcelFormatError("Empty spreadsheet")

        header_row = rows[0]
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(header_row) if cell.value}

        column_map = {
            "id_number": self._settings.column_id_number,
            "email": self._settings.column_email,
            "full_name": self._settings.column_full_name,
            "branch": self._settings.column_branch,
            "phone": self._settings.column_phone,
        }

        required_fields = ["id_number", "email", "full_name", "branch"]
        for field in required_fields:
            col_name = column_map[field]
            if col_name not in headers:
                raise MissingRequiredColumnError(col_name)

        users: list[User] = []
        skipped = 0

        for row_idx, row in enumerate(rows[1:], start=self._settings.header_row + 1):
            raw_data: dict[str, str | None] = {}
            for field, col_name in column_map.items():
                col_idx = headers.get(col_name)
                if col_idx is not None:
                    cell_value = row[col_idx].value
                    raw_data[field] = str(cell_value).strip() if cell_value is not None else None
                else:
                    raw_data[field] = None

            if not raw_data.get("id_number"):
                continue

            try:
                validated = ExcelRowSchema(**raw_data)
            except ValidationError as exc:
                logger.warning("Skipping row %d: %s", row_idx, exc)
                skipped += 1
                continue

            users.append(
                User(
                    id_number=validated.id_number,
                    email=validated.email,
                    full_name=validated.full_name,
                    branch=validated.branch,
                    phone=validated.phone,
                )
            )

        wb.close()

        if not users:
            raise InvalidExcelFormatError("No valid rows found")

        count = await self._user_repo.bulk_upsert(users)
        logger.info("Excel import complete: %d imported, %d skipped", count, skipped)
        return len(users)
